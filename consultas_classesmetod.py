from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import chromedriver_autoinstaller
import json
import openpyxl

class ConfigDriver:
    def driver_config(downloaddir, user_data_dir=None, page_load=False):
        """
        Criação e configuração padrão do webdriver do selenium para Chrome
        Possível passar os parâmetros da configuração rcb acessando como dict
        :param downloaddir: Caminho de download do diretório
        :param driver_path: Caminho do chromedriver.exe
        :return: web_driver
        """
        chromedriver_autoinstaller.install()

        appState = {"recentDestinations": [
            {"id": "Save as PDF", "origin": "local"}], "selectedDestinationId": "Save as PDF", "version": 2}
        options = Options()
        options.add_experimental_option("prefs", {
            "plugins.always_open_pdf_externally": True,
            "download.default_directory": downloaddir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "printing.print_preview_sticky_settings.appState": json.dumps(appState)
        })
        options.add_argument('--kiosk-printing')
        options.add_argument('--no-sandbox')
        options.add_argument('--zoom=70')
        options.add_argument("--start-maximized")
        options.add_argument("--disable-dev-shm-usage")

        if user_data_dir:
            options.add_argument(f'user-data-dir={user_data_dir}')
        if page_load:
            options.page_load_strategy = 'none'

        web_driver = webdriver.Chrome(
            options=options)
        web_driver.implicitly_wait(3)

        return web_driver

class ConfigXlsx():
    """
    Classe para configuração para criação, leitura e manuseio de planilhas xlsx
    Utiliza openpyxl
    Deve-se atentar aos caminhos passados e nas diferenças entre cada ponto
    caminho_completo -> O caminho do arquivo xlsx
    Sheet -> Manuseio geral da planilha, é basicamente o arquivo xlsx
    planilha_principal -> A planilha dentro do arquivo xlsx que está sendo utilizada
    """

    def ler_criar_xlsx(caminho_completo):
        
        """
        Ao passar um caminho e o nome do arquivo xlsx ela verica(lê)/cria(lê)
        :param caminho_completo: Caminho/arquivo.xlsx
        :return: Retorna o arquivo para manuseio (ex: sheet.save)
        """
        try:
            planilha = openpyxl.load_workbook(f'{caminho_completo}')
            print('Arquivo de planilha verificado.')
            return planilha
        except:
            planilha = openpyxl.Workbook()
            planilha.save(f'{caminho_completo}')
            print('Arquivo de planilha criado.')
            return planilha

    def planilha_principal(arquivo_xlsx):
        print('Planilha principal interna criada.')
        return arquivo_xlsx[arquivo_xlsx.sheetnames[0]]

    def criar_cabecalho(lista_cabecalho, planilha_principal, arquivo_principal, caminho_completo):
        """
        Cria um cabeçalho em planilhas novas #ATENÇÂO! RETORNA A 1º LINHA EM BRANCO
        Caso a planilha já possua alguma linha preenchida a função não é realizada
        :param lista_cabecalho: Lista com os nomes que devem encabeçar a planilha
        :param planilha_principal: Planilha na qual o cabeçalho será adicionado
        :param arquivo_principal: Planilha aberta através do Workbook
        :param caminho_completo: Caminho/arquivo.xlsx que deverá ser salvo após a edição
        """
        count = 0
        for lines in planilha_principal:
            count += 1
            break
        if count == 0:
            for column in range(0, len(lista_cabecalho)):
                planilha_principal.cell(1, column+1).value = lista_cabecalho[column]
            print('Cabeçalho criado.')
            arquivo_principal.save(caminho_completo)
        else:
            print('Não foi possível criar cabeçalho. Já há linhas escritas na planilha.')

    def linha_planilha(planilha_principal):
        """
        Função analisa quantas linhas já estão preenchidas na planilha
        :return: Retorna quantidade de linhas+1, sem risco de sobrescrever alguma informação
        """
        linha_planilha = 0
        for linha in planilha_principal:
            linha_planilha += 1
        print(f'Total de linhas: {linha_planilha}. Próxima linha a escrever: {linha_planilha+1}')
        return linha_planilha+1


    def escrever_planilha(planilha_principal, arquivo_principal, caminho_completo, quantidade_colunas, linha_atual, lista_dados):
        for coluna in range(0, quantidade_colunas):
            try:
                planilha_principal.cell(linha_atual, coluna+1).value = lista_dados[coluna]
            except IndexError:
                planilha_principal.cell(linha_atual, coluna+1).value = ' '

        arquivo_principal.save(caminho_completo)


if __name__ == '__main__':
    pc = ConfigXlsx
    planilha_caminho = 'C:\\Users\\adevv\\Desktop\\planilha_teste.xlsx'
    planilha_arquivo = pc.ler_criar_xlsx(planilha_caminho)
    planilha_principal = pc.planilha_principal(planilha_arquivo)
    lista_cabecalho = ['Coluna 1', 'Coluna 2', 'Coluna 3']
    pc.criar_cabecalho(lista_cabecalho, planilha_principal, planilha_arquivo, planilha_caminho)
    linha_atual = pc.linha_planilha(planilha_principal)

    pc.escrever_planilha(planilha_principal, planilha_arquivo, planilha_caminho, 3, linha_atual, ['a', 'b'])
    linha_atual += 1